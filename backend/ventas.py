# backend/ventas.py
from typing import List, Dict, Optional
from datetime import datetime
from sqlalchemy import text
from backend.db import engine
from .productos import adjust_stock, get_product
from .deudas import add_debt
from .clientes import get_client
from .logs import registrar_log
from .utils import generate_id
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from num2words import num2words
from datetime import datetime
import os


# ----------------------------
# Funciones de ventas
# ----------------------------

def list_sales() -> List[Dict]:
    with engine.connect() as conn:
        rows = conn.execute(text("""
            SELECT id, fecha, cliente_id, total, pagado, tipo_pago, productos_vendidos
            FROM ventas
            ORDER BY fecha DESC
        """)).fetchall()
    ventas = []
    for r in rows:
        ventas.append({
            "id": r[0],
            "fecha": r[1].isoformat() if r[1] else None,
            "cliente_id": r[2],
            "total": float(r[3]),
            "pagado": float(r[4]),
            "tipo_pago": r[5],
            "productos_vendidos": r[6]  # ya debe ser JSON en la base
        })
    return ventas

def get_sale(sale_id: str) -> Optional[Dict]:
    with engine.connect() as conn:
        row = conn.execute(
            text("""
                SELECT id, fecha, cliente_id, total, pagado, tipo_pago, productos_vendidos
                FROM ventas
                WHERE id=:id
            """), {"id": sale_id}
        ).fetchone()
        if not row:
            return None
        return {
            "id": row[0],
            "fecha": row[1].isoformat() if row[1] else None,
            "cliente_id": row[2],
            "total": float(row[3]),
            "pagado": float(row[4]),
            "tipo_pago": row[5],
            "productos_vendidos": row[6]
        }

def register_sale(cliente_id: str, items: List[Dict], pagado: float,
                  tipo_pago: Optional[str] = None, usuario: Optional[str] = None) -> Dict:
    cliente = get_client(cliente_id)
    if not cliente:
        raise KeyError(f"Cliente {cliente_id} no existe")

    # Validar stock
    for it in items:
        prod = get_product(it["id_producto"])
        if not prod:
            raise KeyError(f"Producto {it['id_producto']} no existe")
        if prod.get("cantidad", 0) < it["cantidad"]:
            raise ValueError(f"Stock insuficiente para {prod['nombre']} (disponible {prod['cantidad']})")

    # Calcular total
    total = round(sum(float(it["cantidad"]) * float(it["precio_unitario"]) for it in items), 2)

    # Actualizar stock
    for it in items:
        adjust_stock(it["id_producto"], -int(it["cantidad"]))

    # Crear venta
    sale_id = generate_id("V", list_sales())
    fecha = datetime.now()

    with engine.begin() as conn:
        conn.execute(
            text("""
                INSERT INTO ventas (id, fecha, cliente_id, total, pagado, tipo_pago, productos_vendidos)
                VALUES (:id, :fecha, :cliente_id, :total, :pagado, :tipo_pago, :productos_vendidos)
            """),
            {
                "id": sale_id,
                "fecha": fecha,
                "cliente_id": cliente_id,
                "total": total,
                "pagado": pagado,
                "tipo_pago": tipo_pago,
                "productos_vendidos": items  # SQLAlchemy puede guardar JSON directo si la columna es jsonb
            }
        )

    # Registrar deuda si aplica
    if pagado < total:
        add_debt(cliente_id, round(total - pagado, 2), usuario=usuario)

    # Registrar log
    registrar_log(usuario or "sistema", "registrar_venta", {
        "venta_id": sale_id, "cliente_id": cliente_id,
        "total": total, "pagado": pagado, "tipo_pago": tipo_pago,
        "productos": items
    })

    return get_sale(sale_id)

def delete_sale(sale_id: str, usuario: Optional[str] = None) -> bool:
    sale = get_sale(sale_id)
    if not sale:
        return False

    with engine.begin() as conn:
        conn.execute(text("DELETE FROM ventas WHERE id=:id"), {"id": sale_id})

    registrar_log(usuario or "sistema", "eliminar_venta", {"venta_id": sale_id, "venta": sale})
    return True


def generar_factura_excel(venta: dict, cliente: dict, transportista: dict = None, filename: str = None):
    """
    Genera una factura en Excel lista para imprimir.
    - venta: dict con la venta (productos_vendidos, total, pagado, tipo_pago, id, fecha)
    - cliente: dict con información del cliente (nombre, direccion, id, chapa)
    - transportista: dict opcional (nombre, chapa)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"Factura {venta['id']}"

    # ---------------------------
    # Encabezado
    # ---------------------------
    ws.merge_cells("A1:D1")
    ws["A1"] = "ELECTROGALÍNDEZ S.A."
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:D2")
    ws["A2"] = "Calle Falsa 123, La Habana"
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A3:D3")
    ws["A3"] = "NIT: 123456789"
    ws["A3"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A4:D4")
    ws["A4"] = "Actividad: Venta de equipos electrónicos"
    ws["A4"].alignment = Alignment(horizontal="center")

    ws["A6"] = f"Factura Nº: {venta['id']}"
    ws["D6"] = f"Fecha: {venta['fecha'][:10]}"  # YYYY-MM-DD

    # ---------------------------
    # Cliente
    # ---------------------------
    ws["A8"] = f"Cliente: {cliente['nombre']}"
    ws["A9"] = f"Dirección: {cliente.get('direccion', '')}"
    ws["A10"] = f"ID: {cliente.get('id_cliente', cliente.get('id', ''))}"
    ws["A11"] = f"Chapa: {cliente.get('chapa', '')}"

    # ---------------------------
    # Transportista
    # ---------------------------
    if transportista:
        ws["A13"] = f"Transportista: {transportista.get('nombre', '')}"
        ws["A14"] = f"Chapa vehículo: {transportista.get('chapa', '')}"

    # ---------------------------
    # Productos
    # ---------------------------
    ws["A16"] = "Producto"
    ws["B16"] = "Cantidad"
    ws["C16"] = "Precio Unitario"
    ws["D16"] = "Subtotal"
    header_font = Font(bold=True)
    for cell in ["A16", "B16", "C16", "D16"]:
        ws[cell].font = header_font

    row = 17
    for item in venta["productos_vendidos"]:
        ws[f"A{row}"] = item["nombre"]
        ws[f"B{row}"] = item["cantidad"]
        ws[f"C{row}"] = item["precio_unitario"]
        ws[f"D{row}"] = item["cantidad"] * item["precio_unitario"]
        row += 1

    # ---------------------------
    # Totales
    # ---------------------------
    ws[f"C{row}"] = "Total:"
    ws[f"D{row}"] = venta["total"]
    ws[f"C{row+1}"] = "Total en letras:"
    ws[f"D{row+1}"] = num2words(venta["total"], lang="es").capitalize() + " pesos cubanos"
    ws[f"C{row+2}"] = "Forma de pago:"
    ws[f"D{row+2}"] = venta.get("tipo_pago", "")

    # ---------------------------
    # Firma y nota
    # ---------------------------
    ws[f"A{row+4}"] = "Firma del TCP: ____________________________"
    ws[f"A{row+5}"] = "Nota: Factura válida para efectos contables y tributarios"

    # ---------------------------
    # Ajuste de columnas
    # ---------------------------
    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 20

    # ---------------------------
    # Guardar archivo
    # ---------------------------
    if not filename:
        filename = f"factura_{venta['id']}.xlsx"
    wb.save(filename)
    return filename
